#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
股票筛选工具
===========

功能：根据指定条件筛选Excel文件中的股票数据
作者：Manus AI
创建时间：2025-06-08

主要功能：
1. 读取Excel文件中的多个页签数据
2. 根据股票代码进行筛选匹配
3. 生成详细的筛选报告
4. 输出多种格式的结果文件
"""

import os
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


class StockFilter:
    """股票筛选器类"""

    def __init__(self, excel_file_path):
        """
        初始化股票筛选器

        Args:
            excel_file_path (str): Excel文件路径
        """
        self.excel_file_path = excel_file_path
        self.sheet_a_data = None
        self.sheet_b_data = None
        self.filtered_data = None
        self.filter_stats = {}

    def load_data(self, sheet_a_name, sheet_b_name):
        """
        加载Excel文件中的数据

        Args:
            sheet_a_name (str): 页签A名称（待筛选的数据）
            sheet_b_name (str): 页签B名称（参考股票代码）

        Returns:
            bool: 加载是否成功
        """
        try:
            print(f"正在读取Excel文件: {self.excel_file_path}")

            # 读取页签A数据
            self.sheet_a_data = pd.read_excel(self.excel_file_path, sheet_name=sheet_a_name)
            print(f"页签A ({sheet_a_name}) 数据形状: {self.sheet_a_data.shape}")

            # 读取页签B数据
            self.sheet_b_data = pd.read_excel(self.excel_file_path, sheet_name=sheet_b_name)
            print(f"页签B ({sheet_b_name}) 数据形状: {self.sheet_b_data.shape}")

            return True

        except Exception as e:
            print(f"读取Excel文件失败: {e}")
            return False

    def filter_stocks(self, stock_code_column='股票代码'):
        """
        根据股票代码进行筛选

        Args:
            stock_code_column (str): 股票代码列名

        Returns:
            pd.DataFrame: 筛选后的数据
        """
        if self.sheet_a_data is None or self.sheet_b_data is None:
            raise ValueError("请先加载数据")

        print(f"\n开始筛选股票...")

        # 获取页签B中的股票代码集合
        sheet_b_codes = set(self.sheet_b_data[stock_code_column].dropna())
        print(f"页签B中的股票代码数量: {len(sheet_b_codes)}")
        print(f"页签B股票代码: {sorted(list(sheet_b_codes))}")

        # 筛选页签A中股票代码出现在页签B中的记录
        self.filtered_data = self.sheet_a_data[
            self.sheet_a_data[stock_code_column].isin(sheet_b_codes)
        ].copy()

        # 统计筛选结果
        self.filter_stats = {
            'original_count': len(self.sheet_a_data),
            'filtered_count': len(self.filtered_data),
            'filter_ratio': len(self.filtered_data) / len(self.sheet_a_data) * 100,
            'reference_codes_count': len(sheet_b_codes),
            'matched_codes': sorted(list(self.filtered_data[stock_code_column].unique()))
        }

        print(f"\n筛选结果:")
        print(f"筛选前记录数: {self.filter_stats['original_count']}")
        print(f"筛选后记录数: {self.filter_stats['filtered_count']}")
        print(f"筛选比例: {self.filter_stats['filter_ratio']:.2f}%")

        return self.filtered_data

    def display_filtered_stocks(self, display_columns=None):
        """
        显示筛选出的股票信息

        Args:
            display_columns (list): 要显示的列名列表
        """
        if self.filtered_data is None:
            print("请先执行筛选操作")
            return

        if display_columns is None:
            display_columns = ['名称', '股票代码', '现价', '市值(经调整)', 'technical_signal_1d']

        print(f"\n筛选出的股票详情:")
        print("-" * 80)

        for idx, row in self.filtered_data.iterrows():
            stock_info = f"- {row['名称']} (代码: {row['股票代码']})"
            if '现价' in row:
                stock_info += f", 现价: {row['现价']}元"
            if 'technical_signal_1d' in row:
                stock_info += f", 技术信号: {row['technical_signal_1d']}"
            print(stock_info)

    def save_simple_result(self, output_path):
        """
        保存简单的筛选结果

        Args:
            output_path (str): 输出文件路径
        """
        if self.filtered_data is None:
            print("请先执行筛选操作")
            return

        self.filtered_data.to_excel(output_path, index=False, sheet_name='筛选结果')
        print(f"简单筛选结果已保存到: {output_path}")

    def save_detailed_report(self, output_path):
        """
        保存详细的筛选报告

        Args:
            output_path (str): 输出文件路径
        """
        if self.filtered_data is None:
            print("请先执行筛选操作")
            return

        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "筛选结果报告"

        # 设置样式
        title_font = Font(size=16, bold=True)
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # 添加标题
        ws['A1'] = "股票筛选结果报告"
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:F1')

        # 添加生成时间
        ws['A2'] = f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A2:F2')

        # 添加筛选条件说明
        ws['A4'] = "筛选条件："
        ws['A4'].font = header_font
        ws['A5'] = "页签A中股票代码出现在页签B中的股票"

        # 添加统计信息
        ws['A7'] = "筛选统计："
        ws['A7'].font = header_font
        ws['A8'] = f"原始股票数量: {self.filter_stats['original_count']}"
        ws['A9'] = f"筛选后股票数量: {self.filter_stats['filtered_count']}"
        ws['A10'] = f"筛选比例: {self.filter_stats['filter_ratio']:.2f}%"
        ws['A11'] = f"参考代码数量: {self.filter_stats['reference_codes_count']}"

        # 添加数据表头
        headers = ['股票名称', '股票代码', '现价', '市值(经调整)', '1日技术信号', '1周技术信号']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=13, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        # 添加筛选结果数据
        for row_idx, (_, row) in enumerate(self.filtered_data.iterrows(), 14):
            ws.cell(row=row_idx, column=1, value=row.get('名称', ''))
            ws.cell(row=row_idx, column=2, value=row.get('股票代码', ''))
            ws.cell(row=row_idx, column=3, value=row.get('现价', ''))
            ws.cell(row=row_idx, column=4, value=row.get('市值(经调整)', ''))
            ws.cell(row=row_idx, column=5, value=row.get('technical_signal_1d', ''))
            ws.cell(row=row_idx, column=6, value=row.get('technical_signal_1w', ''))

        # 调整列宽
        column_widths = {'A': 15, 'B': 12, 'C': 10, 'D': 15, 'E': 15, 'F': 15}
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # 保存文件
        wb.save(output_path)
        print(f"详细筛选报告已保存到: {output_path}")

    def save_complete_result(self, output_path):
        """
        保存完整的筛选结果（包含所有原始数据）

        Args:
            output_path (str): 输出文件路径
        """
        if self.filtered_data is None:
            print("请先执行筛选操作")
            return

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # 筛选结果
            self.filtered_data.to_excel(writer, sheet_name='筛选结果', index=False)

            # 参考股票代码
            self.sheet_b_data.to_excel(writer, sheet_name='参考股票代码', index=False)

            # 筛选统计
            stats_df = pd.DataFrame([
                ['原始股票数量', self.filter_stats['original_count']],
                ['筛选后股票数量', self.filter_stats['filtered_count']],
                ['筛选比例(%)', f"{self.filter_stats['filter_ratio']:.2f}"],
                ['参考代码数量', self.filter_stats['reference_codes_count']]
            ], columns=['统计项目', '数值'])
            stats_df.to_excel(writer, sheet_name='筛选统计', index=False)

        print(f"完整筛选结果已保存到: {output_path}")

    def save_summary_text(self, output_path):
        """
        保存文本摘要

        Args:
            output_path (str): 输出文件路径
        """
        if self.filtered_data is None:
            print("请先执行筛选操作")
            return

        summary = f"""股票筛选结果摘要
================

筛选时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
筛选条件: 页签A中股票代码出现在页签B中的股票

筛选统计:
- 原始股票数量: {self.filter_stats['original_count']}
- 筛选后股票数量: {self.filter_stats['filtered_count']}
- 筛选比例: {self.filter_stats['filter_ratio']:.2f}%
- 参考代码数量: {self.filter_stats['reference_codes_count']}

筛选出的股票列表:
"""

        for idx, row in self.filtered_data.iterrows():
            stock_info = f"- {row.get('名称', 'N/A')} ({row.get('股票代码', 'N/A')})"
            if '现价' in row and pd.notna(row['现价']):
                stock_info += f": 现价 {row['现价']}元"
            if 'technical_signal_1d' in row and pd.notna(row['technical_signal_1d']):
                stock_info += f", 技术信号 {row['technical_signal_1d']}"
            summary += stock_info + "\n"

        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(summary)

        print(f"筛选结果摘要已保存到: {output_path}")


def main():
    """主函数 - 演示如何使用股票筛选器"""

    # 配置参数
    excel_file = 'cyy - cyy - 2025-06-09.xlsx'  # Excel文件路径
    sheet_a_name = 'cyy - cyy - 2025-06-09'  # 页签A名称
    sheet_b_name = 'Sheet1'  # 页签B名称

    # 输出文件路径
    output_dir = './'
    simple_output = os.path.join(output_dir, 'filtered_stocks.xlsx')
    detailed_output = os.path.join(output_dir, '股票筛选结果报告.xlsx')
    complete_output = os.path.join(output_dir, '完整筛选结果.xlsx')
    summary_output = os.path.join(output_dir, '筛选结果摘要.txt')

    try:
        # 创建筛选器实例
        filter_tool = StockFilter(excel_file)

        # 加载数据
        if not filter_tool.load_data(sheet_a_name, sheet_b_name):
            return

        # 执行筛选
        filtered_result = filter_tool.filter_stocks()

        # 显示筛选结果
        filter_tool.display_filtered_stocks()

        # 保存各种格式的结果
        filter_tool.save_simple_result(simple_output)
        filter_tool.save_detailed_report(detailed_output)
        filter_tool.save_complete_result(complete_output)
        filter_tool.save_summary_text(summary_output)

        print(f"\n所有结果文件已生成完成！")
        print(f"输出目录: {output_dir}")

    except Exception as e:
        print(f"程序执行出错: {e}")


if __name__ == "__main__":
    main()

